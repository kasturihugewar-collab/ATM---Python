from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "ATM_Accounts.xlsx"


# Create Excel file if it does not exist
def create_excel_file():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts"

        ws.append(["Account No.", "Account Holder", "Address", "Balance"])

        wb.save(FILE_NAME)


# Create a new account
def create_account():
    wb = load_workbook(FILE_NAME)
    ws = wb["Accounts"]

    name = input("Enter your name: ")
    address = input("Enter your address: ")

    # Generate account number
    if ws.max_row == 1:
        account_no = 1001
    else:
        account_no = ws.cell(ws.max_row, 1).value + 1

    balance = 1000

    ws.append([account_no, name, address, balance])

    wb.save(FILE_NAME)

    print("\n- ACCOUNT CREATED SUCCESSFULLY -")
    print("Account Number:", account_no)
    print("Account Holder:", name)
    print("Initial Balance: Rs.", balance)


# Find account row
def find_account(account_no):
    wb = load_workbook(FILE_NAME)
    ws = wb["Accounts"]

    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == account_no:
            return row

    return None


# Deposit money
def deposit():
    account_no = int(input("Enter Account Number: "))

    row = find_account(account_no)

    if row is None:
        print("Account not found.")
        return

    amount = float(input("Enter amount to deposit: Rs."))

    if amount > 0:
        wb = load_workbook(FILE_NAME)
        ws = wb["Accounts"]

        current_balance = ws.cell(row, 4).value
        new_balance = current_balance + amount

        ws.cell(row, 4).value = new_balance

        wb.save(FILE_NAME)

        print("Deposited Rs.", amount)
        print("Current Balance: Rs.", new_balance)

    else:
        print("Enter a valid amount.")


# Withdraw money
def withdraw():
    account_no = int(input("Enter Account Number: "))

    row = find_account(account_no)

    if row is None:
        print("Account not found.")
        return

    amount = float(input("Enter amount to withdraw: Rs."))

    wb = load_workbook(FILE_NAME)
    ws = wb["Accounts"]

    current_balance = ws.cell(row, 4).value

    if amount <= 0:
        print("Enter a valid amount.")

    elif current_balance >= amount:
        new_balance = current_balance - amount

        ws.cell(row, 4).value = new_balance

        wb.save(FILE_NAME)

        print("Withdrawal Successful")
        print("Current Balance: Rs.", new_balance)

    else:
        print("Insufficient funds.")
        print("Current Balance: Rs.", current_balance)


# Check balance
def check_balance():
    account_no = int(input("Enter Account Number: "))

    row = find_account(account_no)

    if row is None:
        print("Account not found.")
        return

    wb = load_workbook(FILE_NAME)
    ws = wb["Accounts"]

    name = ws.cell(row, 2).value
    balance = ws.cell(row, 4).value

    print("\nAccount Holder:", name)
    print("Current Balance: Rs.", balance)


# Main program
def main():

    create_excel_file()

    while True:
        print()
        print("====== ATM SYSTEM ======")
        print("1. Create Account")
        print("2. Deposit Money")
        print("3. Withdraw Money")
        print("4. Check Balance")
        print("5. Exit")

        choice = int(input("Choose an option: "))

        if choice == 1:
            create_account()

        elif choice == 2:
            deposit()

        elif choice == 3:
            withdraw()

        elif choice == 4:
            check_balance()

        elif choice == 5:
            print("Thank you for the visit, Goodbye!!")
            break

        else:
            print("Invalid option, Please try again.")


main()
